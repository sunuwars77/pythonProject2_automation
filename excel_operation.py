import os
import openpyxl
from openpyxl.styles import PatternFill, Font

test_result_location = 'TestResult.xlsx'


def write_result(sn, test_summary, result, remarks):
    workbook = openpyxl.load_workbook(test_result_location)
    worksheet2 = workbook.get_sheet_by_name('Details')
    row = int(sn) + 1
    worksheet2.cell(row, 1, sn)
    worksheet2.cell(row, 2, test_summary)
    if result == 'PASS':
        worksheet2.cell(row, 3, result).fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type="solid")
    else:
        worksheet2.cell(row, 3, result).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid")
    worksheet2.cell(row, 4, str(remarks))
    workbook.save(test_result_location)


def write_header():
    workbook = openpyxl.Workbook()
    worksheet2 = workbook.create_sheet('Details')
    worksheet2.cell(1, 1, "SN")
    worksheet2.cell(1, 2, "Test Summary")
    worksheet2.cell(1, 3, "Result")
    worksheet2.cell(1, 4, "Remarks")
    color = "009999FF"
    for rows in worksheet2.iter_rows(min_row=1, max_row=1, min_col=1, max_col=4):
        for cell in rows:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid", fgColor="009999FF")

    workbook.save(test_result_location)


def write_result_summary(test_started, test_completed, url, total_test, passed_test, failed_test, skipped_test):
    workbook = openpyxl.load_workbook(test_result_location)
    worksheet2 = workbook.get_sheet_by_name('Summary')
    worksheet2.cell(1, 2, test_started)
    worksheet2.cell(2, 2, test_completed)
    worksheet2.cell(3, 2, url)
    worksheet2.cell(4, 2, total_test)
    worksheet2.cell(5, 2, passed_test)
    worksheet2.cell(6, 2, failed_test)
    worksheet2.cell(7, 2, skipped_test)
    workbook.save(test_result_location)


def write_header_summary():
    workbook = openpyxl.load_workbook(test_result_location)
    worksheet = workbook.create_sheet('Summary')
    worksheet.cell(1, 1, "Test Started On")
    worksheet.cell(2, 1, "Test Completed On")
    worksheet.cell(3, 1, "Url")
    worksheet.cell(4, 1, "Total Number of Test")
    worksheet.cell(5, 1, "Number of Passed Test Case")
    worksheet.cell(6, 1, "Number of Failed Test Case")
    worksheet.cell(7, 1, "Number of skipped test cases")
    color = "000000FF"
    for rows in worksheet.iter_rows(min_row=1, max_row=7, min_col=1, max_col=1):
        for cell in rows:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid", fgColor="00FFFFFF")
            cell.font = Font(b=True, color="00FFFFFF")
    workbook.save(test_result_location)
